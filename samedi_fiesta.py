from os import path, listdir, rename
import re
from shutil import copyfile
import openpyxl

# path_nro_test = r'C:\Users\boris.petrus\Desktop\test_1\PHOTOS\import\AGT'
# path_nro = r'C:\Users\boris.petrus\Desktop\test_1\PHOTOS\import'
base = r'/home/acid/Pictures/Train_head_sit/'
c6folder = r"/home/acid/martillac/BEinfo/C6/"
c6test = "79BVR_00001_DIS_PRO_ind02_PA 27126_ART9_C6.xlsx"
finish = r"/home/acid/Documents/c6_to_send/"
c6xlsx_test = path.join(c6folder,c6test)
# export = path.join(base,'export')
tete = path.join(base,'tete')
# et = path.join(base,'et')
site = path.join(base,'site')

width = 330
height = 400


def read_picture_name(c6xlsx):
    dicoart = {}
    nro = path.basename(c6xlsx)[2:5]
    sro = path.basename(c6xlsx).split("_")[1]
    pa = path.basename(c6xlsx).split("_")[-3]
    art = path.basename(c6xlsx).split("_")[-2]
    ptname_row = {}
    wb = openpyxl.load_workbook(c6xlsx)
    sheet_photo = wb['Photos']
    for row in range(3, sheet_photo.max_row + 1):
        for column in "A":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if sheet_photo[cell_name].value != 'None':
                pt_name = str(sheet_photo[cell_name].value)
                pt_num = pt_name.split("_")[0]
                ptname_row[pt_num] = row
                # dicoart[art] += pt_num
    return [nro, sro, pa, art], ptname_row

# 
# def rename_photo_by_folder_name(nro):
    folder_photo = path.join(path_nro, nro)
    liste = []
    for folder in listdir(folder_photo):
        if 'FT' in folder:
            foldernew = folder.replace('FT ', "FT")
            actual_folder = path.join(folder_photo, folder)
            pt = re.findall(r'[*FT]\w+', foldernew)[0].replace("FT","")
            if len(pt) < 7:
                pt = "0"+pt
            i = 0
            for img in listdir(actual_folder):
                imgnewname = path.join(base,"{}_{}_unverified_{}".format(nro, pt, i)) + path.splitext(img)[1]
                print(imgnewname)
                copyfile(path.join(actual_folder, img), imgnewname)
                i += 1
            # liste.append()
    return liste


def place_picture(position_path_sit, position_path_tete, c6xlsx):
    wb = openpyxl.load_workbook(path.join(c6folder, c6xlsx))
    sheet_photo = wb['Photos']
    for key, value in position_path_tete.items():
        thumb = openpyxl.drawing.image.Image(value)
        thumb.height = height
        thumb.width = width
        position = "D" + str(int(key)-1)
        sheet_photo.add_image(thumb, position)
    for key, value in position_path_sit.items():
        thumb = openpyxl.drawing.image.Image(value)
        thumb.height = height
        thumb.width = width
        position = "C" + str(int(key)-1)
        sheet_photo.add_image(thumb, position)
    # for row in sheet_photo:
    #     remove(sheet_photo, row)
    wb.save(path.join(c6folder, c6xlsx))
    wb.close()


def add_tete_picture_to_c6(pt_name_row):
    ptnum_tetepath = {}
    export = {}
    for pict in listdir(tete):
        pt_num = pict.split("_")[1]
        ptnum_tetepath[pt_num] = path.join(tete, pict)
    for key in pt_name_row.keys():
        # print(key)
        if key in ptnum_tetepath or "0"+key in ptnum_tetepath:
            export[pt_name_row[key]] = ptnum_tetepath[key]
    # print(export)
    return export


def add_sit_picture_to_c6(pt_name_row):
    # print(dico)
    ptnum_tetepath = {}
    export = {}
    for pict in listdir(site):
        pt_num = pict.split("_")[1]
        ptnum_tetepath[pt_num] = path.join(site, pict)
    for key in pt_name_row.keys():
        # print(key)
        if key in ptnum_tetepath or "0"+key in ptnum_tetepath:
            export[pt_name_row[key]] = ptnum_tetepath[key]
    # print(export)
    return export


def geti_info_c6(vrac):
    Nro = {'nro'}
    Sro = ['sro']
    # Pa = {}
    for c6xlsx in listdir(vrac):
        print(c6xlsx)
        if c6xlsx.endswith(".xlsx"):
            try:
                [nro, sro, pa, art], ptname_row = read_picture_name(path.join(vrac, c6xlsx))
                # print(ptname_row)
                Nro.add(nro)
                Sro.append({nro:sro})
                # print(ptname_row)
                position_path_tete = add_tete_picture_to_c6(ptname_row)
                # print(position_path_tete)
                position_path_ensemble = add_sit_picture_to_c6(ptname_row)
                # place_picture(position_path_ensemble, position_path_tete,path.join(vrac, c6xlsx))
            except:
                print("error on file : {}".format(c6xlsx))

# geti_info_c6(c6folder)

# add_sit_picture_to_c6(path.join(site,"79PRA_00006_DIS_PRO_ind02_V2_PA 26924_ART 06_C6.xlsx"))